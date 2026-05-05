import openpyxl
import sys

def inspect_excel(file_path, log_file):
    log_file.write(f"--- Inspecting {file_path} ---\n")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet in wb.sheetnames:
        log_file.write(f"\nSheet: {sheet}\n")
        ws = wb[sheet]
        
        widths = []
        for col_idx in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            w = ws.column_dimensions[col_letter].width
            widths.append(str(w) if w else "default")
        log_file.write(f"Column Widths: {widths}\n")
        
        # Check header colors (Assuming header is on row 7)
        header_row = 7
        colors = []
        for cell in ws[header_row]:
            fg_color = "None"
            if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color and cell.fill.start_color.rgb:
                fg_color = str(cell.fill.start_color.rgb)
            colors.append(f"{cell.value}: {fg_color}")
        log_file.write(f"Header Colors (Row 8): {colors}\n")

if __name__ == "__main__":
    with open("check_excel_styles_output.txt", "w", encoding="utf-8") as f:
        inspect_excel(sys.argv[1], f)
        inspect_excel(sys.argv[2], f)

