import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys

def format_excel(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    
    # Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def apply_page_setup(ws, orientation):
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = orientation
        # Fit to 1 page wide, auto height
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        # Optional: set some margins to Narrow to fit more data
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

    # ==========================
    # Sheet 1: BOM Chi Tiết (A4 Portrait)
    # ==========================
    if "BOM Chi Tiết" in wb.sheetnames:
        ws = wb["BOM Chi Tiết"]
        apply_page_setup(ws, ws.ORIENTATION_PORTRAIT)
        
        ws['A1'].font = title_font
        ws['A1'].alignment = left_align
        ws['A2'].font = Font(name="Calibri", size=11, italic=True)
        
        # Tăng độ rộng các cột để hiển thị đủ nội dung
        col_widths = {'A': 6, 'B': 22, 'C': 15, 'D': 14, 'E': 14, 'F': 14, 'G': 12, 'H': 16, 'I': 15, 'J': 12}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
            
        header_row = 4
        # Format Header
        for cell in ws[header_row]:
            if type(cell).__name__ != 'MergedCell':
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
        # Format Data
        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
            for cell in row:
                if type(cell).__name__ == 'MergedCell':
                    continue
                cell.font = normal_font
                cell.border = thin_border
                if cell.column_letter in ['A', 'D', 'E', 'G', 'H', 'I', 'J']:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
        # Format Total Row
        if ws.max_row > header_row:
            for cell in ws[ws.max_row]:
                if type(cell).__name__ != 'MergedCell':
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    # ==========================
    # Sheet 2: Hao Hụt (A4 Portrait)
    # ==========================
    if "Hao Hụt" in wb.sheetnames:
        ws = wb["Hao Hụt"]
        apply_page_setup(ws, ws.ORIENTATION_PORTRAIT)
        
        if type(ws['A1']).__name__ != 'MergedCell':
            ws['A1'].font = title_font
        
        col_widths = {'A': 6, 'B': 25, 'C': 15, 'D': 14, 'E': 14, 'F': 16, 'G': 15, 'H': 15, 'I': 12}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
            
        header_row = 3
        for cell in ws[header_row]:
            if type(cell).__name__ != 'MergedCell':
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
            for cell in row:
                if type(cell).__name__ == 'MergedCell':
                    continue
                cell.font = normal_font
                cell.border = thin_border
                # Check column type
                if cell.column_letter in ['A', 'D', 'E', 'F', 'H', 'I']:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
                # Format specific values (e.g. area to 2 decimal places if number)
                try:
                    if cell.column_letter in ['D', 'E', 'F']:
                        val = float(cell.value)
                        cell.value = val
                        if cell.column_letter == 'F':
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '0'
                except:
                    pass

    # ==========================
    # Sheet 3: Tổng Hợp (A4 Landscape)
    # ==========================
    if "Tổng Hợp" in wb.sheetnames:
        ws = wb["Tổng Hợp"]
        apply_page_setup(ws, ws.ORIENTATION_LANDSCAPE)
        
        col_widths = {'A': 40, 'B': 22, 'C': 15, 'D': 15, 'E': 15, 'F': 14, 'G': 18, 'H': 25}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
            
        # TỔNG HỢP DỰ ÁN section
        if type(ws['A1']).__name__ != 'MergedCell':
            ws['A1'].font = Font(name="Calibri", size=12, bold=True, color="000000")
        
        for row in range(3, 15):
            c1 = ws.cell(row=row, column=1)
            c2 = ws.cell(row=row, column=2)
            if type(c1).__name__ != 'MergedCell':
                c1.font = Font(name="Calibri", size=11, bold=True)
            if type(c2).__name__ != 'MergedCell':
                c2.alignment = right_align
            
        if type(ws['A17']).__name__ != 'MergedCell':
            ws['A17'].font = Font(name="Calibri", size=12, bold=True, color="000000")
        
        header_row = 18
        for cell in ws[header_row]:
            if cell.value and type(cell).__name__ != 'MergedCell':
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                
        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
            is_total = (row[0].value == "TỔNG ĐẶT HÀNG")
            for cell in row:
                if type(cell).__name__ == 'MergedCell':
                    continue
                if cell.column <= 8:
                    cell.border = thin_border
                    if is_total:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                    else:
                        cell.font = normal_font
                        
                    if cell.column_letter in ['A', 'C', 'D', 'E', 'F', 'G']:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align

    wb.save(output_path)
    print(f"Formatted Excel saved to: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python format_excel.py <input> <output>")
    else:
        format_excel(sys.argv[1], sys.argv[2])
