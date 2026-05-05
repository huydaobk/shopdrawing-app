import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys

def format_excel(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name="Times New Roman", size=13, bold=True, color="FFFFFF")
    title_font = Font(name="Times New Roman", size=14, bold=True)
    normal_font = Font(name="Times New Roman", size=13)
    meta_font = Font(name="Times New Roman", size=13, italic=True)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align_nowrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def apply_page_setup(ws, orientation):
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = orientation
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

    def insert_project_info(ws, last_col_letter, date_str, rows_to_insert=2):
        import re
        # Remove any existing freeze panes before inserting rows to avoid weird shifts
        ws.freeze_panes = None
        
        # Shift merged cells manually since openpyxl insert_rows does not update them
        old_merges = list(ws.merged_cells.ranges)
        for m in old_merges:
            ws.merged_cells.remove(m)
            
        # Insert rows after the main title (Row 1)
        ws.insert_rows(2, rows_to_insert)
        
        for m in old_merges:
            if m.min_row >= 2:
                m.shift(row_shift=rows_to_insert, col_shift=0)
            ws.merged_cells.add(m)
            
        # Fix formulas that were shifted
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    old_formula = str(cell.value)
                    new_formula = re.sub(
                        r'([A-Z]+)(\d+)', 
                        lambda m: f"{m.group(1)}{int(m.group(2))+rows_to_insert}" if int(m.group(2)) >= 2 else m.group(0), 
                        old_formula
                    )
                    cell.value = new_formula
        
        # Add Project info
        ws['A2'] = "Dự án: [Đồng bộ Tên dự án từ tính năng Input]"
        ws['A3'] = "Địa chỉ: [Đồng bộ Địa chỉ từ tính năng Input]"
        
        if ws.title != "Lệnh Sản Xuất" and date_str:
            ws['A4'] = date_str
        
        # Format the metadata rows (Row 2, 3, and potentially 4 if Ngày xuất is present)
        italic_font = Font(name='Times New Roman', size=13, italic=True)
        align_left_center = Alignment(horizontal='left', vertical='center')
        
        for r in [2, 3, 4]:
            val = ws[f'A{r}'].value
            if val and isinstance(val, str) and ("Dự án:" in val or "Địa chỉ:" in val or "Ngày xuất:" in val):
                ws[f'A{r}'].font = italic_font
                ws[f'A{r}'].alignment = align_left_center
                ws.row_dimensions[r].height = 24  # Increase height to make it less dense (Quy chuẩn khoảng cách dòng)
                
                # Merge the row across all columns
                try:
                    ws.merge_cells(f'A{r}:{last_col_letter}{r}')
                except Exception:
                    pass

    def find_stt_row(ws):
        for i in range(1, 20):
            val = ws[f'A{i}'].value
            if val and str(val).strip().upper() == "STT":
                return i
        return None

    def apply_global_font(ws):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = normal_font

    # Extract global date string before any insertions
    global_date_str = ""
    if "Lệnh Sản Xuất" in wb.sheetnames:
        global_date_str = wb["Lệnh Sản Xuất"]["A2"].value

    # ==========================
    # Sheet 1: Lệnh Sản Xuất
    # ==========================
    if "Lệnh Sản Xuất" in wb.sheetnames:
        ws = wb["Lệnh Sản Xuất"]
        apply_global_font(ws)
        apply_page_setup(ws, ws.ORIENTATION_PORTRAIT)
        
        insert_project_info(ws, 'J', global_date_str, 2)
        
        if type(ws['A1']).__name__ != 'MergedCell':
            ws['A1'].font = title_font
            ws['A1'].alignment = left_align
        ws.row_dimensions[1].height = 30
            
        header_row = find_stt_row(ws)
        if header_row:
            ws.freeze_panes = f"A{header_row + 1}"
            
            col_widths = {'A': 6, 'B': 10, 'C': 12, 'D': 15, 'E': 15, 'F': 14, 'G': 14, 'H': 12, 'I': 16, 'J': 30}
            for col, w in col_widths.items():
                ws.column_dimensions[col].width = w
                
            for cell in ws[header_row]:
                if type(cell).__name__ != 'MergedCell':
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
            for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
                for cell in row:
                    if type(cell).__name__ == 'MergedCell':
                        continue
                    cell.font = normal_font
                    cell.border = thin_border
                    if cell.column_letter in ['A', 'B', 'F', 'G', 'H', 'I']:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
            # Format Total Row
            for i in range(header_row+1, ws.max_row+1):
                val = ws[f'A{i}'].value
                if val and "TỔNG" in str(val).upper():
                    try:
                        ws.merge_cells(f'A{i}:G{i}')
                    except Exception:
                        pass
                    for col in range(1, 11):
                        cell = ws.cell(row=i, column=col)
                        cell.font = Font(name="Times New Roman", size=13, bold=True)
                        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                        cell.border = thin_border
                    ws[f'A{i}'].alignment = center_align_nowrap

    # ==========================
    # Sheet 2: Quản lý Spec
    # ==========================
    if "Quản lý Spec" in wb.sheetnames:
        ws = wb["Quản lý Spec"]
        apply_global_font(ws)
        apply_page_setup(ws, ws.ORIENTATION_PORTRAIT)
        
        insert_project_info(ws, 'R', global_date_str, 2)
        
        if type(ws['A1']).__name__ != 'MergedCell':
            ws['A1'].font = title_font
            ws['A1'].alignment = left_align
        ws.row_dimensions[1].height = 30
            
        header_row = find_stt_row(ws)
        if header_row:
            # Quản lý Spec header spans two rows (e.g. 7 and 8), data starts at header_row + 2
            ws.freeze_panes = f"A{header_row + 2}"
            
            col_widths = {
                'A': 6, 'B': 15, 'C': 15, 'D': 15, 'E': 12, 'F': 15, 'G': 12, 'H': 10,
                'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 12, 'R': 12
            }
            for col, w in col_widths.items():
                ws.column_dimensions[col].width = w
                
            for r in [header_row-1, header_row, header_row+1]:
                if r > 0:
                    for cell in ws[r]:
                        if cell.value and type(cell).__name__ != 'MergedCell':
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_align
                            cell.border = thin_border
                        
            for row in ws.iter_rows(min_row=header_row+2, max_row=ws.max_row):
                for cell in row:
                    if type(cell).__name__ == 'MergedCell':
                        continue
                    if cell.value is not None:
                        cell.font = normal_font
                        cell.border = thin_border
                        cell.alignment = center_align

    # ==========================
    # Sheet 3: Đặt Hàng Phụ Kiện
    # ==========================
    if "Đặt Hàng Phụ Kiện" in wb.sheetnames:
        ws = wb["Đặt Hàng Phụ Kiện"]
        apply_global_font(ws)
        apply_page_setup(ws, ws.ORIENTATION_LANDSCAPE)
        
        insert_project_info(ws, 'H', global_date_str, 3)
        
        if type(ws['A1']).__name__ != 'MergedCell':
            ws['A1'].font = title_font
            ws['A1'].alignment = left_align
        ws.row_dimensions[1].height = 30
            
        header_row = find_stt_row(ws)
        if header_row:
            ws.freeze_panes = f"A{header_row + 1}"
            
            col_widths = {'A': 6, 'B': 15, 'C': 20, 'D': 35, 'E': 30, 'F': 10, 'G': 12, 'H': 40}
            for col, w in col_widths.items():
                ws.column_dimensions[col].width = w
                
            for cell in ws[header_row]:
                if type(cell).__name__ != 'MergedCell':
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
            for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
                for cell in row:
                    if type(cell).__name__ == 'MergedCell':
                        continue
                    if cell.column <= 8:
                        cell.border = thin_border
                        cell.font = normal_font
                        if cell.column_letter in ['A', 'F', 'G']:
                            cell.alignment = center_align
                        else:
                            cell.alignment = left_align
                            
            # Format Total Row
            for i in range(header_row+1, ws.max_row+1):
                val = ws[f'A{i}'].value
                if val and "TỔNG" in str(val).upper():
                    try:
                        ws.merge_cells(f'A{i}:F{i}')
                    except Exception:
                        pass
                    for col in range(1, 9):
                        cell = ws.cell(row=i, column=col)
                        cell.font = Font(name="Times New Roman", size=13, bold=True)
                        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                        cell.border = thin_border
                    ws[f'A{i}'].alignment = center_align_nowrap

    wb.save(output_path)
    print(f"Formatted Excel saved to: {output_path}")

if __name__ == "__main__":
    format_excel(sys.argv[1], sys.argv[2])
