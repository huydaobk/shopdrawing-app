import openpyxl
import sys
import json

def read_excel_structure(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    res = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        data = []
        for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
            row_data = []
            for cell in row:
                row_data.append(str(cell.value) if cell.value is not None else "")
            data.append(row_data)
        
        # also read column widths
        col_widths = {}
        for col_letter, col_dim in ws.column_dimensions.items():
            col_widths[col_letter] = col_dim.width
            
        res[sheet] = {"preview": data, "col_widths": col_widths, "max_row": ws.max_row, "max_col": ws.max_column}
    
    with open("excel_structure.json", "w", encoding="utf-8") as f:
        json.dump(res, f, indent=2, ensure_ascii=False)

if __name__ == "__main__":
    read_excel_structure(sys.argv[1])
